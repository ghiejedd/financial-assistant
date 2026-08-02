"""
Database layer for Financial Assistant.
Async SQLite database manager using aiosqlite.
"""

import psycopg
from psycopg.rows import dict_row
import os
from urllib.parse import urlparse
import os
from datetime import datetime as _dt, timedelta, timezone
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WIB = timezone(timedelta(hours=7))

class datetime(_dt):
    @classmethod
    def now(cls, tz=None):
        if tz is not None:
            return super().now(tz)
        return super().now(WIB).replace(tzinfo=None)

DB_PATH = os.getenv(
    "DB_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "financial_assistant.db")
)
EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")

# Default categories
EXPENSE_CATEGORIES = {
    "makanan": "🍔 Makanan & Minuman",
    "transport": "🚗 Transportasi",
    "belanja": "🛒 Belanja",
    "tagihan": "💡 Tagihan & Utilitas",
    "hiburan": "🎮 Hiburan",
    "kesehatan": "💊 Kesehatan",
    "pendidikan": "📚 Pendidikan",
    "investasi": "💰 Investasi",
    "lainnya": "📦 Lainnya",
}

INCOME_CATEGORIES = {
    "gaji": "💼 Gaji",
    "freelance": "💻 Freelance",
    "investasi": "📈 Hasil Investasi",
    "bonus": "🎁 Bonus",
    "lainnya": "📦 Lainnya",
}

# Keyword mapping for auto-categorization
CATEGORY_KEYWORDS = {
    "makanan": [
        "makan", "minum", "nasi", "kopi", "coffee", "snack", "jajan",
        "sarapan", "lunch", "dinner", "breakfast", "resto", "warung",
        "bakso", "sate", "ayam", "noodle", "mie", "pizza", "burger",
        "es", "teh", "juice", "boba", "starbucks", "kfc", "mcd",
        "gopay food", "grabfood", "shopeefood", "gofood", "cemilan",
    ],
    "transport": [
        "gojek", "grab", "uber", "taxi", "taksi", "ojek", "bensin",
        "bbm", "parkir", "tol", "bus", "kereta", "krl", "mrt", "lrt",
        "angkot", "transjakarta", "ojol", "travel", "tiket",
    ],
    "belanja": [
        "shopee", "tokped", "tokopedia", "lazada", "blibli", "beli",
        "belanja", "baju", "celana", "sepatu", "tas", "gadget",
        "elektronik", "mall", "supermarket", "indomaret", "alfamart",
    ],
    "tagihan": [
        "listrik", "air", "pdam", "wifi", "internet", "pulsa", "paket data",
        "telkom", "indihome", "sewa", "kos", "kontrakan", "rent",
        "cicilan", "kredit", "asuransi", "pajak", "iuran",
    ],
    "hiburan": [
        "game", "netflix", "spotify", "youtube", "bioskop", "cinema",
        "nonton", "main", "wisata", "liburan", "hangout", "karaoke",
        "disney", "steam", "playstation", "xbox", "subscription",
    ],
    "kesehatan": [
        "obat", "dokter", "rumah sakit", "rs", "klinik", "apotek",
        "vitamin", "supplement", "gym", "fitness", "olahraga",
    ],
    "pendidikan": [
        "buku", "kursus", "course", "udemy", "les", "sekolah",
        "kuliah", "spp", "seminar", "training", "workshop",
    ],
    "investasi": [
        "saham", "reksadana", "crypto", "bitcoin", "nabung", "deposito",
        "obligasi", "trading", "invest",
    ],
    "gaji": [
        "gaji", "salary", "payroll", "upah",
    ],
    "freelance": [
        "freelance", "project", "proyek", "client", "klien", "jasa",
    ],
    "bonus": [
        "bonus", "thr", "reward", "hadiah", "cashback",
    ],
}


async def init_db():
    """Initialize the database and create tables if they don't exist."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id SERIAL PRIMARY KEY,
                telegram_user_id INTEGER NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('income', 'expense')),
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS budgets (
                id SERIAL PRIMARY KEY,
                telegram_user_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                monthly_limit REAL NOT NULL,
                UNIQUE(telegram_user_id, category)
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS savings_goals (
                id SERIAL PRIMARY KEY,
                telegram_user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                target_amount REAL NOT NULL,
                current_amount REAL DEFAULT 0,
                icon TEXT DEFAULT '🎯',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS savings_transactions (
                id SERIAL PRIMARY KEY,
                goal_id INTEGER NOT NULL REFERENCES savings_goals(id),
                telegram_user_id INTEGER NOT NULL,
                amount REAL NOT NULL,
                note TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS accounts (
                id SERIAL PRIMARY KEY,
                telegram_user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                account_type TEXT DEFAULT 'bank',
                balance REAL DEFAULT 0,
                icon TEXT DEFAULT '💳',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(telegram_user_id, name)
            )
        """)
        await db.execute("""
            CREATE INDEX IF NOT EXISTS idx_transactions_user
            ON transactions(telegram_user_id, created_at DESC)
        """)
        try:
            await db.execute("ALTER TABLE transactions ADD COLUMN account_name TEXT")
        except Exception:
            pass
        try:
            await db.execute("ALTER TABLE savings_goals ADD COLUMN goal_type TEXT DEFAULT 'personal'")
        except Exception:
            pass
        await db.commit()
    
    os.makedirs(EXPORTS_DIR, exist_ok=True)


async def detect_account_in_text(user_id: int, text: str) -> Optional[str]:
    """
    Detect account name mentioned in text (e.g. 'BRI', 'BSI', 'BCA', 'Mandiri', 'GoPay', 'Dana', etc.).
    Returns matched account name (string) or None.
    """
    import re
    text_clean = text.lower().strip()
    words = re.findall(r'\b[a-z0-9_]+\b', text_clean)

    # 1. Match against existing accounts in DB for this user
    user_accounts = await get_accounts(user_id)
    for acc in user_accounts:
        acc_name_lower = acc["name"].lower().strip()
        if acc_name_lower in words or acc_name_lower in text_clean:
            return acc["name"]

    # 2. Match against known bank / e-wallet keywords
    for kw in ACCOUNT_ICONS.keys():
        if kw in ("bank", "ewallet", "investasi"):
            continue
        if kw in words or kw in text_clean:
            if len(kw) <= 4:
                return kw.upper()
            return kw.capitalize()

    return None


async def add_transaction(
    user_id: int,
    tx_type: str,
    category: str,
    amount: float,
    description: str = "",
    account_name: Optional[str] = None,
) -> dict:
    """Add a new transaction and return it, updating account balance if account_name is provided."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            INSERT INTO transactions (telegram_user_id, type, category, amount, description, account_name, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (user_id, tx_type, category, amount, description, account_name, datetime.now().isoformat()),
        )
        await db.commit()
        tx_id = cursor.lastrowid

    account_info = None
    if account_name:
        # Get current balance of this account
        user_accounts = await get_accounts(user_id)
        current_acc = next((a for a in user_accounts if a["name"].lower() == account_name.lower()), None)
        current_balance = current_acc["balance"] if current_acc else 0.0

        delta = amount if tx_type == "income" else -amount
        new_balance = current_balance + delta

        # Update account in DB
        acc_type = current_acc["account_type"] if current_acc else ("bank" if account_name.upper() in ["BRI", "BSI", "BCA", "MANDIRI", "BNI", "CIMB"] else "ewallet")
        acc_data = await add_or_update_account(
            user_id=user_id,
            name=account_name,
            balance=new_balance,
            account_type=acc_type,
        )
        account_info = {
            "name": acc_data["name"],
            "balance": new_balance,
            "delta": delta,
            "icon": acc_data["icon"],
        }

    return {
        "id": tx_id,
        "telegram_user_id": user_id,
        "type": tx_type,
        "category": category,
        "amount": amount,
        "description": description,
        "account_name": account_name,
        "account_info": account_info,
        "created_at": datetime.now().isoformat(),
    }


async def delete_last_transaction(user_id: int) -> Optional[dict]:
    """Delete the most recent transaction for a user and revert account balance."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            SELECT * FROM transactions
            WHERE telegram_user_id = %s
            ORDER BY created_at DESC LIMIT 1
            """,
            (user_id,),
        )
        row = await cursor.fetchone()
        if not row:
            return None

        tx = dict(row)
        await db.execute("DELETE FROM transactions WHERE id = %s", (tx["id"],))
        await db.commit()

    if tx.get("account_name"):
        acc_name = tx["account_name"]
        user_accounts = await get_accounts(user_id)
        current_acc = next((a for a in user_accounts if a["name"].lower() == acc_name.lower()), None)
        if current_acc:
            reverse_delta = -tx["amount"] if tx["type"] == "income" else tx["amount"]
            new_balance = current_acc["balance"] + reverse_delta
            await add_or_update_account(user_id, acc_name, new_balance, current_acc["account_type"])

    return tx


async def delete_transaction_by_id(user_id: int, tx_id: int) -> Optional[dict]:
    """Delete a specific transaction by ID and revert account balance."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM transactions WHERE id = %s AND telegram_user_id = %s",
            (tx_id, user_id),
        )
        row = await cursor.fetchone()
        if not row:
            return None

        tx = dict(row)
        await db.execute("DELETE FROM transactions WHERE id = %s", (tx_id,))
        await db.commit()

    if tx.get("account_name"):
        acc_name = tx["account_name"]
        user_accounts = await get_accounts(user_id)
        current_acc = next((a for a in user_accounts if a["name"].lower() == acc_name.lower()), None)
        if current_acc:
            reverse_delta = -tx["amount"] if tx["type"] == "income" else tx["amount"]
            new_balance = current_acc["balance"] + reverse_delta
            await add_or_update_account(user_id, acc_name, new_balance, current_acc["account_type"])

    return tx


async def edit_transaction(
    user_id: int,
    tx_id: int,
    new_amount: Optional[float] = None,
    new_description: Optional[str] = None,
    new_type: Optional[str] = None,
    new_category: Optional[str] = None,
    new_account_name: Optional[str] = None,
    new_created_at: Optional[str] = None,
) -> Optional[dict]:
    """Edit a transaction's details. Adjusts account balances if needed."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM transactions WHERE id = %s AND telegram_user_id = %s",
            (tx_id, user_id),
        )
        row = await cursor.fetchone()
        if not row:
            return None

        tx = dict(row)
        old_amount = tx["amount"]
        old_type = tx["type"]
        old_account = tx.get("account_name")

        updated_amount = new_amount if new_amount is not None else old_amount
        updated_description = new_description if new_description is not None else tx["description"]
        updated_type = new_type if new_type is not None else old_type
        updated_category = new_category if new_category is not None else tx["category"]
        
        # If explicitly passed empty string, it removes the account
        if new_account_name is not None:
            updated_account = new_account_name if new_account_name.strip() else None
        else:
            updated_account = old_account

        updated_created_at = new_created_at if new_created_at is not None else tx["created_at"]

        await db.execute(
            "UPDATE transactions SET amount = %s, description = %s, type = %s, category = %s, account_name = %s, created_at = %s WHERE id = %s",
            (updated_amount, updated_description, updated_type, updated_category, updated_account, updated_created_at, tx_id),
        )
        await db.commit()

    # Adjust account balances
    if old_account or updated_account:
        user_accounts = await get_accounts(user_id)
        
        # Revert old transaction from old account
        if old_account:
            acc = next((a for a in user_accounts if a["name"].lower() == old_account.lower()), None)
            if acc:
                revert_delta = old_amount if old_type == "expense" else -old_amount
                acc["balance"] += revert_delta
                await add_or_update_account(user_id, old_account, acc["balance"], acc["account_type"])

        # Apply new transaction to new account
        if updated_account:
            # Re-fetch in case old_account was the same as updated_account
            user_accounts = await get_accounts(user_id)
            acc = next((a for a in user_accounts if a["name"].lower() == updated_account.lower()), None)
            if acc:
                apply_delta = updated_amount if updated_type == "income" else -updated_amount
                acc["balance"] += apply_delta
                await add_or_update_account(user_id, updated_account, acc["balance"], acc["account_type"])

    tx["amount"] = updated_amount
    tx["description"] = updated_description
    tx["type"] = updated_type
    tx["category"] = updated_category
    tx["account_name"] = updated_account
    tx["created_at"] = updated_created_at
    return tx

async def get_transactions(
    user_id: int,
    limit: int = 20,
    offset: int = 0,
    tx_type: Optional[str] = None,
    days: Optional[int] = None,
) -> list[dict]:
    """Get transactions for a user with optional filters."""
    query = "SELECT * FROM transactions WHERE telegram_user_id = %s"
    params: list = [user_id]

    if tx_type:
        query += " AND type = %s"
        params.append(tx_type)

    if days:
        since = (datetime.now() - timedelta(days=days)).isoformat()
        query += " AND created_at >= %s"
        params.append(since)

    query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
    params.extend([limit, offset])

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(query, params)
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def get_summary(user_id: int, days: int = 30) -> dict:
    """Get financial summary for a user."""
    since = (datetime.now() - timedelta(days=days)).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        # Total income
        cursor = await db.execute(
            """
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE telegram_user_id = %s AND type = 'income' AND created_at >= %s
            """,
            (user_id, since),
        )
        total_income = (await cursor.fetchone())[0]

        # Total expense
        cursor = await db.execute(
            """
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense' AND created_at >= %s
            """,
            (user_id, since),
        )
        total_expense = (await cursor.fetchone())[0]

        # Transaction count
        cursor = await db.execute(
            """
            SELECT COUNT(*) FROM transactions
            WHERE telegram_user_id = %s AND created_at >= %s
            """,
            (user_id, since),
        )
        tx_count = (await cursor.fetchone())[0]

    balance = total_income - total_expense
    savings_rate = (balance / total_income * 100) if total_income > 0 else 0

    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "savings_rate": round(savings_rate, 1),
        "transaction_count": tx_count,
        "period_days": days,
    }


async def get_today_spending(user_id: int) -> dict:
    """Get today's spending breakdown."""
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:

        cursor = await db.execute(
            """
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense' AND created_at >= %s
            GROUP BY category
            ORDER BY total DESC
            """,
            (user_id, today_start),
        )
        rows = await cursor.fetchall()

        categories = [dict(row) for row in rows]
        total = sum(c["total"] for c in categories)

    return {"total": total, "categories": categories}


async def get_daily_spending(user_id: int, days: int = 30) -> list[dict]:
    """Get daily spending for the last N days."""
    since = (datetime.now() - timedelta(days=days)).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            SELECT DATE(created_at) as date,
                   SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expense,
                   SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as income
            FROM transactions
            WHERE telegram_user_id = %s AND created_at >= %s
            GROUP BY DATE(created_at)
            ORDER BY date ASC
            """,
            (user_id, since),
        )
        rows = await cursor.fetchall()

        # Fill gaps with zero
        result = []
        date_data = {row[0]: {"expense": row[1], "income": row[2]} for row in rows}

        for i in range(days):
            date = (datetime.now() - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
            data = date_data.get(date, {"expense": 0, "income": 0})
            result.append({
                "date": date,
                "expense": data["expense"],
                "income": data["income"],
            })

        return result


async def get_category_breakdown(user_id: int, days: int = 30) -> list[dict]:
    """Get spending breakdown by category."""
    since = (datetime.now() - timedelta(days=days)).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense' AND created_at >= %s
            GROUP BY category
            ORDER BY total DESC
            """,
            (user_id, since),
        )
        rows = await cursor.fetchall()
        return [{"category": row[0], "total": row[1], "count": row[2]} for row in rows]


async def get_monthly_trend(user_id: int, months: int = 6) -> list[dict]:
    """Get monthly income vs expense trend."""
    since = (datetime.now() - timedelta(days=months * 30)).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            SELECT strftime('%Y-%m', created_at) as month,
                   SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as income,
                   SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expense
            FROM transactions
            WHERE telegram_user_id = %s AND created_at >= %s
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month ASC
            """,
            (user_id, since),
        )
        rows = await cursor.fetchall()
        return [{"month": row[0], "income": row[1], "expense": row[2]} for row in rows]


async def get_all_user_ids() -> list[int]:
    """Get all unique user IDs from the database."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT DISTINCT telegram_user_id FROM transactions"
        )
        rows = await cursor.fetchall()
        return [row[0] for row in rows]


async def export_to_excel(user_id: int, days: Optional[int] = None) -> str:
    """Export transactions to Excel file and return the file path."""
    transactions = await get_transactions(user_id, limit=10000, days=days)
    summary = await get_summary(user_id, days=days or 30)

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    filename = f"financial_report_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = Workbook()

    # ── Summary Sheet ──
    ws_summary = wb.active
    ws_summary.title = "Ringkasan"

    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    label_font = Font(name="Calibri", bold=True, size=11)
    value_font = Font(name="Calibri", size=11)
    money_format = '#,##0'

    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "💰 Financial Assistant — Ringkasan Keuangan"
    ws_summary["A1"].font = header_font
    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    summary_data = [
        ("", ""),
        ("📅 Periode", f"{summary['period_days']} hari terakhir"),
        ("💵 Total Pemasukan", summary["total_income"]),
        ("💸 Total Pengeluaran", summary["total_expense"]),
        ("💰 Saldo", summary["balance"]),
        ("📊 Savings Rate", f"{summary['savings_rate']}%"),
        ("📝 Jumlah Transaksi", summary["transaction_count"]),
    ]

    for i, (label, value) in enumerate(summary_data, start=2):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = label_font
        ws_summary[f"B{i}"] = value
        ws_summary[f"B{i}"].font = value_font
        if isinstance(value, (int, float)):
            ws_summary[f"B{i}"].number_format = money_format

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # ── Transactions Sheet ──
    ws_tx = wb.create_sheet("Transaksi")

    headers = ["No", "Tanggal", "Tipe", "Kategori", "Jumlah (Rp)", "Deskripsi"]
    header_fill_tx = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    header_font_tx = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, start=1):
        cell = ws_tx.cell(row=1, column=col, value=header)
        cell.font = header_font_tx
        cell.fill = header_fill_tx
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    income_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    expense_fill = PatternFill(start_color="ffe4e6", end_color="ffe4e6", fill_type="solid")

    for i, tx in enumerate(transactions, start=2):
        fill = income_fill if tx["type"] == "income" else expense_fill
        row_data = [
            i - 1,
            tx["created_at"][:19].replace("T", " "),
            "Pemasukan" if tx["type"] == "income" else "Pengeluaran",
            tx["category"],
            tx["amount"],
            tx["description"],
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws_tx.cell(row=i, column=col, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col == 5:
                cell.number_format = money_format

    ws_tx.column_dimensions["A"].width = 6
    ws_tx.column_dimensions["B"].width = 20
    ws_tx.column_dimensions["C"].width = 14
    ws_tx.column_dimensions["D"].width = 22
    ws_tx.column_dimensions["E"].width = 18
    ws_tx.column_dimensions["F"].width = 30

    wb.save(filepath)
    return filepath


def auto_categorize(description: str) -> tuple[str, str]:
    """
    Auto-categorize a transaction based on description keywords.
    Returns (category_key, category_display_name).
    """
    desc_lower = description.lower()

    for cat_key, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in desc_lower:
                # Determine if it's income or expense category
                if cat_key in INCOME_CATEGORIES:
                    return cat_key, INCOME_CATEGORIES[cat_key]
                elif cat_key in EXPENSE_CATEGORIES:
                    return cat_key, EXPENSE_CATEGORIES[cat_key]

    return "lainnya", EXPENSE_CATEGORIES["lainnya"]


# ══════════════════════════════════════════════
# Savings Goals
# ══════════════════════════════════════════════

SAVINGS_ICONS = {
    "darurat": "🛡️", "emergency": "🛡️",
    "liburan": "✈️", "vacation": "✈️", "travel": "✈️",
    "nikah": "💍", "wedding": "💍",
    "rumah": "🏠", "house": "🏠", "kos": "🏠",
    "mobil": "🚗", "motor": "🏍️", "car": "🚗",
    "gadget": "📱", "laptop": "💻", "hp": "📱",
    "pendidikan": "🎓", "kuliah": "🎓", "sekolah": "🎓",
    "investasi": "📈", "invest": "📈", "saham": "📈",
    "kesehatan": "💊", "health": "💊",
}


def pick_savings_icon(name: str) -> str:
    """Pick an appropriate icon for a savings goal name."""
    name_lower = name.lower()
    for keyword, icon in SAVINGS_ICONS.items():
        if keyword in name_lower:
            return icon
    return "🎯"


async def add_savings_goal(
    user_id: int, name: str, target_amount: float, goal_type: str = "personal"
) -> dict:
    """Create a new savings goal."""
    icon = pick_savings_icon(name)
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            """
            INSERT INTO savings_goals (telegram_user_id, name, target_amount, icon, created_at, goal_type)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (user_id, name, target_amount, icon, datetime.now().isoformat(), goal_type),
        )
        await db.commit()
        return {
            "id": cursor.lastrowid,
            "name": name,
            "target_amount": target_amount,
            "current_amount": 0,
            "icon": icon,
            "progress": 0,
            "goal_type": goal_type,
        }


async def deposit_savings(
    user_id: int, goal_name: str, amount: float, note: str = ""
) -> Optional[dict]:
    """Deposit money into a savings goal."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM savings_goals WHERE telegram_user_id = %s AND LOWER(name) = LOWER(%s)",
            (user_id, goal_name),
        )
        goal = await cursor.fetchone()
        if not goal:
            return None

        goal = dict(goal)
        new_amount = goal["current_amount"] + amount

        await db.execute(
            "UPDATE savings_goals SET current_amount = %s WHERE id = %s",
            (new_amount, goal["id"]),
        )
        await db.execute(
            """
            INSERT INTO savings_transactions (goal_id, telegram_user_id, amount, note, created_at)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (goal["id"], user_id, amount, note or f"Setor ke {goal_name}", datetime.now().isoformat()),
        )
        await db.commit()

        goal["current_amount"] = new_amount
        goal["progress"] = round(new_amount / goal["target_amount"] * 100, 1) if goal["target_amount"] > 0 else 0
        return goal


async def withdraw_savings(
    user_id: int, goal_name: str, amount: float, note: str = ""
) -> Optional[dict]:
    """Withdraw money from a savings goal."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM savings_goals WHERE telegram_user_id = %s AND LOWER(name) = LOWER(%s)",
            (user_id, goal_name),
        )
        goal = await cursor.fetchone()
        if not goal:
            return None

        goal = dict(goal)
        new_amount = max(0, goal["current_amount"] - amount)

        await db.execute(
            "UPDATE savings_goals SET current_amount = %s WHERE id = %s",
            (new_amount, goal["id"]),
        )
        await db.execute(
            """
            INSERT INTO savings_transactions (goal_id, telegram_user_id, amount, note, created_at)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (goal["id"], user_id, -amount, note or f"Tarik dari {goal_name}", datetime.now().isoformat()),
        )
        await db.commit()

        goal["current_amount"] = new_amount
        goal["progress"] = round(new_amount / goal["target_amount"] * 100, 1) if goal["target_amount"] > 0 else 0
        return goal


async def get_savings_goals(user_id: int) -> list[dict]:
    """Get all savings goals for a user."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM savings_goals WHERE telegram_user_id = %s ORDER BY created_at ASC",
            (user_id,),
        )
        rows = await cursor.fetchall()
        goals = []
        for row in rows:
            g = dict(row)
            g["progress"] = round(g["current_amount"] / g["target_amount"] * 100, 1) if g["target_amount"] > 0 else 0
            goals.append(g)
        return goals


async def delete_savings_goal(user_id: int, goal_name: str) -> bool:
    """Delete a savings goal by name."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "DELETE FROM savings_goals WHERE telegram_user_id = %s AND LOWER(name) = LOWER(%s)",
            (user_id, goal_name),
        )
        await db.commit()
        return cursor.rowcount > 0


# ══════════════════════════════════════════════
# Budget Management
# ══════════════════════════════════════════════

async def set_budget(user_id: int, category: str, monthly_limit: float) -> dict:
    """Set or update a monthly budget for a category."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        await db.execute(
            """
            INSERT INTO budgets (telegram_user_id, category, monthly_limit)
            VALUES (%s, %s, %s)
            ON CONFLICT(telegram_user_id, category)
            DO UPDATE SET monthly_limit = excluded.monthly_limit
            """,
            (user_id, category, monthly_limit),
        )
        await db.commit()
    return {"category": category, "monthly_limit": monthly_limit}


async def get_budgets(user_id: int) -> list[dict]:
    """Get all budget configurations for a user."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM budgets WHERE telegram_user_id = %s ORDER BY category ASC",
            (user_id,),
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def get_budget_vs_actual(user_id: int) -> list[dict]:
    """Compare budget limits vs actual spending for current month."""
    now = datetime.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        # Get budgets
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            "SELECT * FROM budgets WHERE telegram_user_id = %s",
            (user_id,),
        )
        budgets = {row["category"]: dict(row) for row in await cursor.fetchall()}

        # Get actual spending per category this month
        cursor = await db.execute(
            """
            SELECT category, SUM(amount) as spent, COUNT(*) as count
            FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense' AND created_at >= %s
            GROUP BY category
            """,
            (user_id, month_start),
        )
        actuals = {row["category"]: dict(row) for row in await cursor.fetchall()}

    result = []
    # Process budgeted categories
    all_categories = set(list(budgets.keys()) + list(actuals.keys()))
    for cat in sorted(all_categories):
        budget = budgets.get(cat, {})
        actual = actuals.get(cat, {})
        limit = budget.get("monthly_limit", 0)
        spent = actual.get("spent", 0)
        pct = round(spent / limit * 100, 1) if limit > 0 else 0

        status = "safe"
        if limit > 0:
            if pct >= 100:
                status = "over"
            elif pct >= 80:
                status = "warning"

        result.append({
            "category": cat,
            "budget": limit,
            "spent": spent,
            "remaining": max(0, limit - spent),
            "percentage": pct,
            "status": status,
            "has_budget": limit > 0,
        })

    # Sort: budgeted first, then by percentage descending
    result.sort(key=lambda x: (-int(x["has_budget"]), -x["percentage"]))
    return result


# ══════════════════════════════════════════════
# Behavior Analysis & Recommendations
# ══════════════════════════════════════════════

async def get_behavior_analysis(user_id: int) -> dict:
    """Analyze spending behavior and generate recommendations."""
    now = datetime.now()
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        # This month spending by category
        cursor = await db.execute(
            """
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense' AND created_at >= %s
            GROUP BY category ORDER BY total DESC
            """,
            (user_id, this_month_start.isoformat()),
        )
        this_month = {row[0]: {"total": row[1], "count": row[2]} for row in await cursor.fetchall()}

        # Last month spending by category
        cursor = await db.execute(
            """
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM transactions
            WHERE telegram_user_id = %s AND type = 'expense'
              AND created_at >= %s AND created_at < %s
            GROUP BY category ORDER BY total DESC
            """,
            (user_id, last_month_start.isoformat(), this_month_start.isoformat()),
        )
        last_month = {row[0]: {"total": row[1], "count": row[2]} for row in await cursor.fetchall()}

        # This month totals
        cursor = await db.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN type='income' THEN amount END), 0) as income,
                COALESCE(SUM(CASE WHEN type='expense' THEN amount END), 0) as expense
            FROM transactions
            WHERE telegram_user_id = %s AND created_at >= %s
            """,
            (user_id, this_month_start.isoformat()),
        )
        row = await cursor.fetchone()
        total_income = row[0]
        total_expense = row[1]

        # Last month totals
        cursor = await db.execute(
            """
            SELECT COALESCE(SUM(CASE WHEN type='expense' THEN amount END), 0)
            FROM transactions
            WHERE telegram_user_id = %s AND created_at >= %s AND created_at < %s
            """,
            (user_id, last_month_start.isoformat(), this_month_start.isoformat()),
        )
        last_month_total_expense = (await cursor.fetchone())[0]

        # Daily average this month
        days_elapsed = max(1, now.day)
        daily_avg = total_expense / days_elapsed

        # Budget data
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            "SELECT * FROM budgets WHERE telegram_user_id = %s",
            (user_id,),
        )
        budgets = {row["category"]: dict(row) for row in await cursor.fetchall()}

    # ── Calculate Spending Score (0-100) ──
    score = 100
    deductions = []

    # Savings rate factor (40 points max)
    if total_income > 0:
        savings_rate = (total_income - total_expense) / total_income
        if savings_rate >= 0.3:
            pass  # Perfect
        elif savings_rate >= 0.2:
            score -= 10
        elif savings_rate >= 0.1:
            score -= 20
        elif savings_rate >= 0:
            score -= 30
        else:
            score -= 40
            deductions.append("Pengeluaran melebihi pemasukan!")

    # Budget adherence (30 points max)
    over_budget_count = 0
    for cat, data in this_month.items():
        budget = budgets.get(cat, {})
        limit = budget.get("monthly_limit", 0)
        if limit > 0 and data["total"] > limit:
            over_budget_count += 1
    if over_budget_count > 0:
        score -= min(30, over_budget_count * 10)

    # Month-over-month trend (30 points max)
    if last_month_total_expense > 0:
        change = (total_expense - last_month_total_expense) / last_month_total_expense
        if change > 0.3:
            score -= 30
        elif change > 0.15:
            score -= 20
        elif change > 0:
            score -= 10

    score = max(0, min(100, score))

    # ── Top 3 Categories ──
    top_categories = []
    for cat, data in sorted(this_month.items(), key=lambda x: x[1]["total"], reverse=True)[:3]:
        last = last_month.get(cat, {"total": 0})
        change_pct = 0
        if last["total"] > 0:
            change_pct = round((data["total"] - last["total"]) / last["total"] * 100, 1)

        top_categories.append({
            "category": cat,
            "amount": data["total"],
            "count": data["count"],
            "change_pct": change_pct,
            "trend": "up" if change_pct > 5 else "down" if change_pct < -5 else "stable",
        })

    # ── Generate Recommendations ──
    recommendations = []

    # Budget-based recommendations
    for cat, data in this_month.items():
        budget = budgets.get(cat, {})
        limit = budget.get("monthly_limit", 0)
        if limit > 0:
            pct = data["total"] / limit * 100
            if pct >= 100:
                over_amount = data["total"] - limit
                recommendations.append({
                    "type": "danger",
                    "icon": "🔴",
                    "title": f"{cat} Over Budget!",
                    "message": f"Sudah melebihi budget Rp {limit:,.0f} sebesar Rp {over_amount:,.0f}. Kurangi pengeluaran di kategori ini.".replace(",", "."),
                })
            elif pct >= 80:
                remaining = limit - data["total"]
                recommendations.append({
                    "type": "warning",
                    "icon": "🟡",
                    "title": f"{cat} Hampir Limit",
                    "message": f"Sudah {pct:.0f}% dari budget. Sisa Rp {remaining:,.0f}.".replace(",", "."),
                })

    # Trend-based recommendations
    for cat, data in this_month.items():
        last = last_month.get(cat, {"total": 0})
        if last["total"] > 0:
            change = (data["total"] - last["total"]) / last["total"] * 100
            if change > 25:
                recommendations.append({
                    "type": "warning",
                    "icon": "📈",
                    "title": f"{cat} Naik {change:.0f}%",
                    "message": f"Pengeluaran naik dari bulan lalu. Evaluasi kebutuhan di kategori ini.",
                })

    # Savings rate recommendation
    if total_income > 0:
        sr = (total_income - total_expense) / total_income * 100
        if sr >= 30:
            recommendations.append({
                "type": "success",
                "icon": "🌟",
                "title": f"Savings Rate {sr:.0f}% — Excellent!",
                "message": "Kamu hemat banget bulan ini. Pertahankan!",
            })
        elif sr >= 20:
            recommendations.append({
                "type": "success",
                "icon": "✅",
                "title": f"Savings Rate {sr:.0f}% — Good",
                "message": "Sudah bagus! Coba tingkatkan ke 30%.",
            })
        elif sr < 10:
            recommendations.append({
                "type": "danger",
                "icon": "⚠️",
                "title": f"Savings Rate {sr:.0f}% — Rendah",
                "message": "Coba kurangi pengeluaran non-esensial untuk meningkatkan tabungan.",
            })

    # Daily average insight
    if daily_avg > 0:
        projected_monthly = daily_avg * 30
        recommendations.append({
            "type": "info",
            "icon": "📊",
            "title": f"Rata-rata Harian: Rp {daily_avg:,.0f}".replace(",", "."),
            "message": f"Proyeksi pengeluaran bulan ini: Rp {projected_monthly:,.0f}.".replace(",", "."),
        })

    # Overall trend
    overall_trend = "stable"
    overall_change = 0
    if last_month_total_expense > 0:
        overall_change = round(
            (total_expense - last_month_total_expense) / last_month_total_expense * 100, 1
        )
        overall_trend = "up" if overall_change > 5 else "down" if overall_change < -5 else "stable"

    return {
        "score": score,
        "total_income": total_income,
        "total_expense": total_expense,
        "daily_average": round(daily_avg),
        "overall_trend": overall_trend,
        "overall_change": overall_change,
        "top_categories": top_categories,
        "recommendations": recommendations,
        "over_budget_count": over_budget_count,
    }


async def get_trend_data(user_id: int, period: str = "daily") -> list[dict]:
    """Get trend data based on period: daily, monthly, or annual."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        if period == "daily":
            return await get_daily_spending(user_id, days=30)

        elif period == "monthly":
            since = (datetime.now() - timedelta(days=365)).isoformat()
            cursor = await db.execute(
                """
                SELECT strftime('%Y-%m', created_at) as period,
                       SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expense,
                       SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as income
                FROM transactions
                WHERE telegram_user_id = %s AND created_at >= %s
                GROUP BY strftime('%Y-%m', created_at)
                ORDER BY period ASC
                """,
                (user_id, since),
            )
            rows = await cursor.fetchall()
            return [{"date": r[0], "expense": r[1], "income": r[2]} for r in rows]

        elif period == "annual":
            cursor = await db.execute(
                """
                SELECT strftime('%Y', created_at) as period,
                       SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expense,
                       SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as income
                FROM transactions
                WHERE telegram_user_id = %s
                GROUP BY strftime('%Y', created_at)
                ORDER BY period ASC
                """,
                (user_id,),
            )
            rows = await cursor.fetchall()
            return [{"date": r[0], "expense": r[1], "income": r[2]} for r in rows]

    return []


# ══════════════════════════════════════════════
# Bank & E-Wallet Account Management
# ══════════════════════════════════════════════

ACCOUNT_ICONS = {
    "bca": "🏦", "mandiri": "🏦", "bni": "🏦", "bri": "🏦", "cimb": "🏦",
    "sebank": "🏦", "jago": "🏦", "blu": "🏦", "jenius": "🏦", "bank": "🏦",
    "gopay": "📱", "ovo": "📱", "dana": "📱", "linkaja": "📱", "shopeepay": "📱", "ewallet": "📱",
    "cash": "💵", "tunai": "💵", "dompet": "💵",
    "bibit": "📈", "bareksa": "📈", "crypto": "🪙", "stockbit": "📈", "investasi": "📈",
}


def pick_account_icon(name: str) -> str:
    """Pick an icon for an account based on name."""
    name_lower = name.lower()
    for kw, icon in ACCOUNT_ICONS.items():
        if kw in name_lower:
            return icon
    return "💳"


async def add_or_update_account(
    user_id: int, name: str, balance: float, account_type: str = "bank"
) -> dict:
    """Add or update an account balance."""
    icon = pick_account_icon(name)
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        await db.execute(
            """
            INSERT INTO accounts (telegram_user_id, name, account_type, balance, icon, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT(telegram_user_id, name)
            DO UPDATE SET balance = excluded.balance, account_type = excluded.account_type, icon = excluded.icon
            """,
            (user_id, name, account_type, balance, icon, datetime.now().isoformat()),
        )
        await db.commit()
    return {"name": name, "balance": balance, "account_type": account_type, "icon": icon}


async def get_accounts(user_id: int) -> list[dict]:
    """Get all accounts and total balance for a user."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "SELECT * FROM accounts WHERE telegram_user_id = %s ORDER BY balance DESC",
            (user_id,),
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def delete_account(user_id: int, name: str) -> bool:
    """Delete an account by name."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute(
            "DELETE FROM accounts WHERE telegram_user_id = %s AND LOWER(name) = LOWER(%s)",
            (user_id, name),
        )
        await db.commit()
        return cursor.rowcount > 0


async def toggle_transaction_type(tx_id: int) -> Optional[dict]:
    """Toggle transaction type between income and expense and adjust account balance."""
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        db_url = "postgresql://neondb_owner:npg_TcbPujgh81ty@ep-noisy-haze-axb17xuz.c-4.us-east-2.aws.neon.tech/neondb?sslmode=require"
    async with await psycopg.AsyncConnection.connect(db_url, row_factory=dict_row) as db:
        cursor = await db.execute("SELECT * FROM transactions WHERE id = %s", (tx_id,))
        row = await cursor.fetchone()
        if not row:
            return None
        tx = dict(row)
        old_type = tx["type"]
        new_type = "income" if old_type == "expense" else "expense"

        # Update category display
        new_cat = tx["category"]
        if old_type == "expense" and new_cat in EXPENSE_CATEGORIES.values():
            new_cat = INCOME_CATEGORIES.get("lainnya", "📦 Lainnya")
        elif old_type == "income" and new_cat in INCOME_CATEGORIES.values():
            new_cat = EXPENSE_CATEGORIES.get("lainnya", "📦 Lainnya")

        await db.execute(
            "UPDATE transactions SET type = %s, category = %s WHERE id = %s",
            (new_type, new_cat, tx_id),
        )
        await db.commit()
        tx["type"] = new_type
        tx["category"] = new_cat

    # Adjust account balance if linked
    account_info = None
    account_name = tx.get("account_name")
    if account_name:
        user_id = tx["telegram_user_id"]
        user_accounts = await get_accounts(user_id)
        current_acc = next((a for a in user_accounts if a["name"].lower() == account_name.lower()), None)
        if current_acc:
            balance_diff = (2 * tx["amount"]) if new_type == "income" else (-2 * tx["amount"])
            new_balance = current_acc["balance"] + balance_diff
            acc_data = await add_or_update_account(
                user_id=user_id,
                name=account_name,
                balance=new_balance,
                account_type=current_acc["account_type"],
            )
            account_info = {
                "name": acc_data["name"],
                "balance": new_balance,
                "icon": acc_data["icon"],
            }
    tx["account_info"] = account_info
    return tx
